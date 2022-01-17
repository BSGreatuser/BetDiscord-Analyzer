from openpyxl.styles import PatternFill, Font
from openpyxl import load_workbook, Workbook
import time, sys, os

print('https://discord.gg/PQ99eZdBBS')
if not os.path.isfile('기록.txt'):
    print('기록.txt 파일이 없습니다.')
    time.sleep(5)
    sys.exit(1)
if os.path.isfile('log.xlsx'):
    os.remove('log.xlsx')
 
Workbook().save('log.xlsx')

with open('기록.txt', encoding='UTF8') as f:
    lines = f.readlines()

if '회차정보\n' in lines:
    lines.remove('회차정보\n')
    with open('기록.txt', 'w', encoding='UTF8') as f:
        f.write(''.join(lines))

x = 1
y = 1

for line in lines:
    wb = load_workbook('log.xlsx')
    ws = wb.active
    xlfile = 'log.xlsx'
    ft = Font(color="FFFFFF")

    color_odd = PatternFill(start_color='0000FF', end_color='0000FF', fill_type='solid')  # 홀
    color_even = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')  # 짝

    spt = line.split(" ")
    nowRound = int(spt[0].replace('회차', ''))
    nowPick = spt[1].strip()

    postion = lines.index(line)
    if postion == 0:
        lastLine = lines[postion]
    else:
        lastLine = lines[postion - 1]

    lastRound = int(lastLine.split(" ")[0].replace('회차', ''))
    lastPick = lastLine.split(" ")[1].strip()

    if nowRound == 1:
        x = x
        y = y
    elif nowPick == lastPick:
        x += 1
    elif nowPick != lastPick:
        y += 1
        x = 1

    c = ws.cell(x, y)
    if nowPick == '홀':
        c.fill = color_odd
    elif nowPick == '짝':
        c.fill = color_even
    c.value = nowRound
    c.font = ft
    wb.save(xlfile)

print("끝")
time.sleep(5)
sys.exit()
